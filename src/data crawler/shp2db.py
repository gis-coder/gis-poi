#!/usr/bin/env python3.7
# -*- coding:utf-8 -*-

import os
import osgeo
from openpyxl import load_workbook
from openpyxl import Workbook
import gis.shpfile as gis_shp
import db.postgre.operate as post_db


def shp2db(shp_path):
    try:
        shp = gis_shp.ShapeRead(shp_path)
        # print(shp.geoType)
        # print(shp.spatial)
        # fields = shp.read_field()
        # print(fields)
        attrs = shp.read_values()
        if attrs is None:
            raise Exception('读取属性表失败')
        geo_path = os.path.dirname(os.path.dirname(os.path.dirname(os.getcwd()))) + '\\data\\region\\geom.txt'
        file = open(geo_path, 'w', encoding='utf-8')
        for row in attrs:
            sql = "update tb_region set f_geom = {0}('{1}',4326) where f_code = '{2}';" \
                .format(__get_func(row[3]), row[3], row[1])
            # post_db.run(sql)
            file.write(sql + "\n")
        file.close()
    except Exception as e:
        print(e)


def __get_func(wkt):
    if wkt.find('MULTIPOLYGON') != -1:
        return 'st_multipolygonfromtext'
    elif wkt.find('MULTILINESTRING') != -1:
        return 'st_multilinestringfromtext'
    elif wkt.find('MULTIPOINT') != -1:
        return 'st_multipointfromtext'
    elif wkt.find('POLYGON') != -1:
        return 'st_polygonfromtext'
    elif wkt.find('LINESTRING') != -1:
        return 'st_linefromtext'
    elif wkt.find('POINT') != -1:
        return 'st_pointfromtext'
    return 'st_geogfromtext'


# 将网上下载的行政区信息导入数据库
def xls2db(xls_path):
    try:
        workbook = load_workbook(xls_path)
        if workbook is None:
            raise Exception('读取电子表格文件失败')
        region_path = os.path.dirname(os.path.dirname(os.path.dirname(os.getcwd()))) + '\\data\\region\\region.txt'
        file = open(region_path, 'w', encoding='utf-8')
        sheets = ['p', 'c', 'd']
        for sheet in sheets:
            worksheet = workbook[sheet]
            if worksheet is None:
                continue
            for row in worksheet.iter_rows(2):
                line = [col.value for col in row]
                sql = "insert into tb_region(f_name,f_code,f_level) values('{0}','{1}','{2}');".format(
                    line[1], line[0], line[2]
                )
                post_db.run(sql)
                file.write(sql + "\n")
        file.close()
    except Exception as e:
        print(e)


if __name__ == '__main__':
    try:
        xls_path = os.path.dirname(
            os.path.dirname(os.path.dirname(os.getcwd()))) + '\\data\\region\\region.xlsx'
        shp_path = os.path.dirname(
            os.path.dirname(os.path.dirname(os.getcwd()))) + '\\data\\region\\polygon\\province.shp'
        # xls2db(xls_path)
        shp2db(shp_path)
        print('Complate')
    except Exception as e:
        print(e)
