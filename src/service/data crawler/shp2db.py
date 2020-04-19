#!/usr/bin/env python3.7
# -*- coding:utf-8 -*-

import osgeo
from openpyxl import load_workbook
from openpyxl import Workbook
import gis.shpfile
import db.postgre.operate as postdb


def shp2db(shp_path):
    try:
        shp = gis.shpfile.ShapeRead(shp_path)
        # print(shp.geoType)
        # print(shp.spatial)
        # fields = shp.read_field()
        # print(fields)
        attrs = shp.read_values()
        if attrs is None:
            raise Exception('读取属性表失败')
        for row in attrs:
            sql = "update tb_region set geom = ST_GeomFromText('{0}',4326) where f_code = '{1}';".format(row[3], row[1])
            postdb.run(sql)
    except Exception as e:
        print(e)


# 将网上下载的行政区信息导入数据库
def xls2db(xls_path):
    try:
        workbook = load_workbook(xls_path)
        if workbook is None:
            raise Exception('读取电子表格文件失败')
        file = open(r'D:\Code\gis-poi\data\region\region.txt', 'w', encoding='utf-8')
        sheets = ['p', 'c', 'd']
        for sheet in sheets:
            worksheet = workbook[sheet]
            if worksheet is None:
                continue
            for row in worksheet.iter_rows(2):
                line = [col.value for col in row]
                sql = "insert into tb_region(f_name,f_code,f_level) values('{0}','{1}','{2}');".format(
                    line[1], line[0], line[2]
                )
                file.write(sql + "\n")
        file.close()
    except Exception as e:
        print(e)


if __name__ == '__main__':
    try:
        shp_path = r'D:\Code\gis-poi\data\region\polygon'
        xls_path = r'D:\Code\gis-poi\data\region\region.xlsx'
        # xls2db(xls_path)
        shp2db(shp_path + '\province.shp')
        print('Complate')
    except Exception as e:
        print(e)
